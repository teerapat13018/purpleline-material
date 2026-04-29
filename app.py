import streamlit as st
import pandas as pd
import gspread
import requests as _req
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook
import os
import io
import shutil
import tempfile
from datetime import datetime

# ──────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

FORM_TEMPLATE_PATH = os.path.join(BASE_DIR, "Form เบิกของ TAM 19032026.xlsx")
SIGNATURE_PATH     = os.path.join(BASE_DIR, "signature.png")
SIGNATURE_PASSWORD = "TAM2026"

SHEET_ID = "1d6uZQdtOCLiWo4EbJ0lg_UPqn7m8BFxyXZ4U5Zg2yXM"
SCOPES   = ["https://www.googleapis.com/auth/spreadsheets"]

LOCATION_OPTIONS = ["— ไม่ระบุ —", "General", "PP25", "VS02NB", "VS02SB", "VS03"]

st.set_page_config(
    page_title="ระบบเบิกวัสดุ Purple Line",
    page_icon="🔧",
    layout="wide",
)

# ──────────────────────────────────────────
# GSPREAD CONNECTION
# ──────────────────────────────────────────

@st.cache_resource
def get_gsheet():
    creds = Credentials.from_service_account_info(
        st.secrets["gcp_service_account"],
        scopes=SCOPES,
    )
    gc = gspread.authorize(creds)
    return gc.open_by_key(SHEET_ID)


def upload_temp(data: bytes, filename: str) -> str:
    """ลองอัปโหลดทีละ service จนสำเร็จ"""
    mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    errors = []

    # 1) tmpfiles.org
    try:
        r = _req.post("https://tmpfiles.org/api/v1/upload",
                      files={"file": (filename, io.BytesIO(data), mime)},
                      timeout=20)
        r.raise_for_status()
        j = r.json()
        if j.get("status") == "success":
            return j["data"]["url"].replace("tmpfiles.org/", "tmpfiles.org/dl/")
        errors.append(f"tmpfiles={j}")
    except Exception as e:
        errors.append(f"tmpfiles={e}")

    # 2) 0x0.st
    try:
        r = _req.post("https://0x0.st/",
                      files={"file": (filename, io.BytesIO(data), mime)},
                      timeout=20)
        if r.ok and r.text.strip().startswith("http"):
            return r.text.strip()
        errors.append(f"0x0=status{r.status_code}:{r.text[:80]}")
    except Exception as e:
        errors.append(f"0x0={e}")

    # 3) litterbox.catbox.moe
    try:
        r = _req.post(
            "https://litterbox.catbox.moe/resources/internals/api.php",
            data={"reqtype": "fileupload", "time": "1h"},
            files={"fileToUpload": (filename, io.BytesIO(data), mime)},
            timeout=30,
        )
        if r.ok and r.text.strip().startswith("http"):
            return r.text.strip()
        errors.append(f"litterbox=status{r.status_code}:{r.text[:80]}")
    except Exception as e:
        errors.append(f"litterbox={e}")

    raise RuntimeError(" | ".join(errors))


# ──────────────────────────────────────────
# DATA LOADERS
# ──────────────────────────────────────────

@st.cache_data(ttl=3600, show_spinner=False)
def load_employees():
    sh = get_gsheet()
    return sh.worksheet("employees").get_all_records()  # [{"name":..., "id":...}]


@st.cache_data(ttl=3600, show_spinner=False)
def load_costcodes():
    sh = get_gsheet()
    records = sh.worksheet("costcodes").get_all_records()
    return pd.DataFrame(records)


@st.cache_data(ttl=300, show_spinner="กำลังโหลดข้อมูล Stock Card...")
def load_stock_data():
    sh = get_gsheet()
    records = sh.worksheet("stock").get_all_records()
    if not records:
        return pd.DataFrame()
    df = pd.DataFrame(records)
    df["คงเหลือ"] = pd.to_numeric(df["คงเหลือ"], errors="coerce").fillna(0)
    df["_key"] = df.index.astype(str)
    return df


# ──────────────────────────────────────────
# DOCUMENT NUMBER  (นับจาก history sheet)
# ──────────────────────────────────────────

def _count_today_docs():
    prefix = "MRQ-104-TAM-" + datetime.now().strftime("%y-%m-%d")
    try:
        sh = get_gsheet()
        col = sh.worksheet("history").col_values(1)  # column A = เลขที่เอกสาร
        return sum(1 for v in col[1:] if v.startswith(prefix))
    except Exception:
        return 0


def peek_next_doc_number() -> str:
    date_str = datetime.now().strftime("%y-%m-%d")
    n = _count_today_docs() + 1
    return f"MRQ-104-TAM-{date_str}-{n:03d}"


def get_next_doc_number() -> str:
    return peek_next_doc_number()


# ──────────────────────────────────────────
# HISTORY  (Google Sheets)
# ──────────────────────────────────────────

def _load_history_raw():
    try:
        sh = get_gsheet()
        return sh.worksheet("history").get_all_records()
    except Exception:
        return []


def load_history() -> pd.DataFrame:
    records = _load_history_raw()
    if not records:
        return pd.DataFrame()
    df = pd.DataFrame(records)
    return df.iloc[::-1].reset_index(drop=True)  # ใหม่สุดขึ้นก่อน


def save_history(selected_items, doc_number: str, requester_name: str = "", employee_id: str = ""):
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    rows = [
        [
            doc_number, now_str, requester_name, employee_id,
            it["รายละเอียด"], float(it["จำนวน"]), it["หน่วย"], it.get("หมายเหตุ", ""),
        ]
        for it in selected_items
    ]
    if rows:
        sh = get_gsheet()
        sh.worksheet("history").append_rows(rows, value_input_option="USER_ENTERED")


def delete_all_history():
    sh = get_gsheet()
    ws = sh.worksheet("history")
    last_row = ws.row_count
    if last_row > 1:
        ws.delete_rows(2, last_row)


def history_to_excel_bytes() -> bytes:
    df = load_history()
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="ประวัติการเบิก")
    buf.seek(0)
    return buf.getvalue()


# ──────────────────────────────────────────
# EXPORT TO FORM
# ──────────────────────────────────────────

def _get_safe_form_path():
    tmp = os.path.join(tempfile.gettempdir(), "form_template_tmp.xlsx")
    try:
        shutil.copy2(FORM_TEMPLATE_PATH, tmp)
        return tmp
    except Exception:
        return FORM_TEMPLATE_PATH


def export_to_form(
    selected_items,
    doc_number: str,
    requester_name: str = "",
    employee_id: str = "",
    costcode: str = "",
    location: str = "",
):
    wb = load_workbook(_get_safe_form_path())
    ws = wb.active
    today_str = datetime.now().strftime("%d/%m/%Y")

    name_value = requester_name.strip()
    if employee_id.strip():
        name_value = f"{name_value}  ({employee_id.strip()})" if name_value else employee_id.strip()
    ws.cell(row=5, column=4).value  = name_value
    ws.cell(row=5, column=10).value = doc_number
    ws.cell(row=7, column=10).value = today_str
    if location:
        ws.cell(row=38, column=2).value = location

    for idx, item in enumerate(selected_items):
        row = 10 + idx
        if row > 35:
            break
        ws.cell(row=row, column=2).value = idx + 1
        ws.cell(row=row, column=3).value = ""
        ws.cell(row=row, column=4).value = item["รายละเอียด"]
        ws.cell(row=row, column=5).value = item["จำนวน"]
        ws.cell(row=row, column=6).value = item["หน่วย"]
        ws.cell(row=row, column=7).value = item.get("หมายเหตุ", "")
        ws.cell(row=row, column=9).value = costcode
        ws.cell(row=row, column=10).value = ""

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ──────────────────────────────────────────
# LOAD DATA (top-level — cached)
# ──────────────────────────────────────────

EMPLOYEES      = load_employees()
COSTCODE_TABLE = load_costcodes()
EMPLOYEE_LABELS = ["— กรุณาเลือกผู้เบิก —"] + [
    f"{e['name']}  ({e['id']})" for e in EMPLOYEES
]
df = load_stock_data()

# ──────────────────────────────────────────
# MAIN UI
# ──────────────────────────────────────────

st.markdown("""
<style>
@media (max-width: 768px) {
    section[data-testid="stSidebar"] { min-width: 0px !important; }
    button[data-baseweb="tab"] { font-size: 15px !important; padding: 10px 8px !important; }
    input[type="number"], input[type="text"] { font-size: 16px !important; min-height: 44px !important; }
    .stButton > button { min-height: 48px !important; font-size: 16px !important; }
    div[data-testid="column"] { width: 100% !important; flex: none !important; }
}
.item-card {
    background: #f8f9fa; border-radius: 10px;
    padding: 12px 14px; margin-bottom: 10px; border: 1px solid #dee2e6;
}
</style>
""", unsafe_allow_html=True)

st.title("🔧 ระบบเลือกและเบิกวัสดุ")

if df.empty:
    st.error("❌ ไม่พบข้อมูล กรุณาตรวจสอบ Google Sheet")
    st.stop()

# ── SESSION STATE ──
for key, default in [
    ("selected_keys", set()),
    ("dl_bytes", None),
    ("dl_fname", None),
    ("dl_msg", None),
    ("dl_drive_url", None),
    ("dl_upload_err", None),
    ("pending_history", None),
]:
    if key not in st.session_state:
        st.session_state[key] = default


def _save_pending_history():
    ph = st.session_state.get("pending_history")
    if ph:
        save_history(ph["qty_data"], ph["doc_number"], ph["requester_name"], ph["employee_id"])
        st.session_state.pending_history = None


# ── SIDEBAR ──
with st.sidebar:
    st.header("⚙️ ตั้งค่า")
    st.caption(f"รายการทั้งหมด: {len(df):,}")
    st.caption(f"เลือกแล้ว: {len(st.session_state.selected_keys)} รายการ")
    st.divider()
    if st.button("🗑️ ล้างรายการที่เลือกทั้งหมด", use_container_width=True):
        st.session_state.selected_keys = set()
        st.rerun()
    if st.button("🔄 รีเฟรชข้อมูล (Stock)", use_container_width=True):
        st.cache_data.clear()
        st.rerun()

# ── DOWNLOAD BAR (เหนือ tabs — แสดงตลอดหลังสร้าง Form) ──
if st.session_state.dl_bytes is not None:
    _save_pending_history()

    # อัปโหลดครั้งแรกเท่านั้น (dl_drive_url == "" หมายความว่าลองแล้วล้มเหลวทุก service)
    if st.session_state.dl_drive_url is None:
        with st.spinner("⏳ กำลังเตรียมไฟล์..."):
            try:
                url = upload_temp(st.session_state.dl_bytes, st.session_state.dl_fname)
                st.session_state.dl_drive_url = url
            except Exception as e:
                st.session_state.dl_drive_url = ""   # mark ว่าล้มเหลว
                st.session_state.dl_upload_err = str(e)
        st.rerun()

    dl_col, close_col = st.columns([5, 1])
    with dl_col:
        if st.session_state.dl_drive_url:
            # อัปโหลดสำเร็จ — ใช้ external link
            st.link_button(
                label=f"⬇️ ดาวน์โหลด {st.session_state.dl_fname}",
                url=st.session_state.dl_drive_url,
                use_container_width=True,
                type="primary",
            )
        else:
            # fallback — ใช้ st.download_button
            st.warning(f"⚠️ upload ไม่สำเร็จ ({st.session_state.get('dl_upload_err','')[:120]}) — ลองกดปุ่มนี้แทน:")
            st.download_button(
                label=f"⬇️ ดาวน์โหลด {st.session_state.dl_fname}",
                data=st.session_state.dl_bytes,
                file_name=st.session_state.dl_fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )
    with close_col:
        if st.button("✖", use_container_width=True, help="ปิด"):
            st.session_state.dl_bytes      = None
            st.session_state.dl_fname      = None
            st.session_state.dl_msg        = None
            st.session_state.dl_drive_url  = None
            st.session_state.dl_upload_err = None
            st.rerun()
    if st.session_state.dl_msg:
        st.caption(st.session_state.dl_msg)
    st.divider()

# ── 3 TABS ──
n_selected = len(st.session_state.selected_keys)
tab1, tab2, tab3 = st.tabs([
    "🔍 ค้นหา & เลือก",
    f"🛒 รายการที่เลือก ({n_selected})",
    "📜 ประวัติ",
])

# ══════════════════════════════════════════
# TAB 1 — ค้นหา & เลือก
# ══════════════════════════════════════════
with tab1:
    col_search, col_avail = st.columns([5, 2])
    with col_search:
        search_text = st.text_input("🔍 ค้นหารายการ", placeholder="พิมพ์ชื่ออุปกรณ์...")
    with col_avail:
        st.write("")
        show_avail = st.checkbox("เฉพาะที่มีคงเหลือ", value=False)

    filtered = df.copy()
    if search_text:
        filtered = filtered[filtered["รายละเอียด"].str.contains(search_text, case=False, na=False)]
    if show_avail:
        filtered = filtered[filtered["คงเหลือ"] > 0]
    filtered = filtered.reset_index(drop=True)

    MAX_SHOW = 50
    total    = len(filtered)
    show_df  = filtered.head(MAX_SHOW)

    if total == 0:
        st.info("ไม่พบรายการ — ลองพิมพ์คำค้นหาใหม่")
    else:
        if total > MAX_SHOW:
            st.caption(f"แสดง {MAX_SHOW} จาก {total:,} รายการ — พิมพ์เพิ่มเพื่อกรองให้แคบลง")
        else:
            st.caption(f"แสดง {total:,} รายการ")

        for _, row in show_df.iterrows():
            k      = row["_key"]
            is_sel = k in st.session_state.selected_keys
            c_btn, c_info = st.columns([1, 4])
            with c_btn:
                if is_sel:
                    if st.button("✅ เพิ่มแล้ว", key=f"add_{k}", use_container_width=True):
                        st.session_state.selected_keys.discard(k)
                        st.rerun()
                else:
                    if st.button("➕ เพิ่ม", key=f"add_{k}", use_container_width=True):
                        st.session_state.selected_keys.add(k)
                        st.rerun()
            with c_info:
                st.markdown(f"**{row['รายละเอียด']}**")
                meta = []
                if row.get("หน่วย"):
                    meta.append(row["หน่วย"])
                if row.get("คงเหลือ") is not None:
                    meta.append(f"คงเหลือ {row['คงเหลือ']:.0f}")
                if meta:
                    st.caption(" · ".join(meta))

    if n_selected:
        st.success(f"เลือกแล้ว {n_selected} รายการ — ไปที่แท็บ 🛒 เพื่อกรอกข้อมูลและ Export")


# ══════════════════════════════════════════
# TAB 2 — รายการที่เลือก + Export
# ══════════════════════════════════════════
with tab2:
    sel_rows = df[df["_key"].isin(st.session_state.selected_keys)].copy()

    if sel_rows.empty:
        st.info("ยังไม่ได้เลือกรายการ — กลับไปแท็บ 🔍 แล้วติ๊ก ✅ รายการที่ต้องการ")
    else:
        qty_data = []

        st.markdown("#### 📦 รายการที่เลือก")
        for _, row in sel_rows.iterrows():
            with st.container():
                st.markdown(f"**{row['รายละเอียด']}**")
                meta = []
                if row["หน่วย"]:
                    meta.append(f"หน่วย: {row['หน่วย']}")
                meta.append(f"คงเหลือ: {row['คงเหลือ']:.0f}")
                if row.get("sheet"):
                    meta.append(f"📁 {row['sheet']}")
                st.caption(" | ".join(meta))

                c_qty, c_remark = st.columns([1, 3])
                with c_qty:
                    qty = st.number_input(
                        "จำนวน", min_value=0.0, value=1.0, step=1.0,
                        key=f"qty_{row['_key']}",
                    )
                with c_remark:
                    remark_item = st.text_input(
                        "ใช้งานอะไร ที่ไหน",
                        key=f"remark_{row['_key']}",
                        placeholder="เช่น ซ่อมปั๊ม VS02SB...",
                    )

                qty_data.append({
                    "รายละเอียด": row["รายละเอียด"],
                    "หน่วย":      row["หน่วย"],
                    "จำนวน":      qty,
                    "หมายเหตุ":   remark_item,
                })
                st.divider()

        # ── ข้อมูลผู้เบิก ──
        st.markdown("#### 👤 ข้อมูลผู้เบิก")
        col_emp, col_docnum = st.columns(2)

        with col_emp:
            emp_choice = st.selectbox("เลือกผู้เบิก", options=EMPLOYEE_LABELS, key="emp_choice")
            if emp_choice == EMPLOYEE_LABELS[0]:
                requester_name = ""
                employee_id    = ""
            else:
                idx            = EMPLOYEE_LABELS.index(emp_choice) - 1
                requester_name = EMPLOYEES[idx]["name"]
                employee_id    = EMPLOYEES[idx]["id"]
            if requester_name:
                st.caption(f"รหัส: **{employee_id}**")

        with col_docnum:
            auto_doc         = peek_next_doc_number()
            doc_number_input = st.text_input(
                "เลขที่เอกสาร (แก้ไขได้)", value=auto_doc, key="doc_number_input",
            )

        # ── Cost Code ──
        st.markdown("#### 💰 Cost Code")
        cc_col1, cc_col2, cc_col3 = st.columns(3)
        with cc_col1:
            cc_type_opts = ["— ประเภท —"] + sorted(COSTCODE_TABLE["ประเภท"].unique().tolist())
            cc_type = st.selectbox("ประเภท", options=cc_type_opts, key="cc_type")
        with cc_col2:
            if cc_type != "— ประเภท —":
                work_opts = ["— Work —"] + sorted(
                    COSTCODE_TABLE[COSTCODE_TABLE["ประเภท"] == cc_type]["Work"].unique().tolist()
                )
            else:
                work_opts = ["— Work —"] + sorted(COSTCODE_TABLE["Work"].unique().tolist())
            cc_work = st.selectbox("Work", options=work_opts, key="cc_work")
        with cc_col3:
            mask = pd.Series([True] * len(COSTCODE_TABLE))
            if cc_type != "— ประเภท —":
                mask &= COSTCODE_TABLE["ประเภท"] == cc_type
            if cc_work != "— Work —":
                mask &= COSTCODE_TABLE["Work"] == cc_work
            loc_opts = ["— Location —"] + sorted(COSTCODE_TABLE[mask]["Location"].unique().tolist())
            cc_loc = st.selectbox("Location", options=loc_opts, key="cc_loc")

        cc_match = COSTCODE_TABLE[
            (COSTCODE_TABLE["ประเภท"]   == cc_type) &
            (COSTCODE_TABLE["Work"]      == cc_work) &
            (COSTCODE_TABLE["Location"]  == cc_loc)
        ]
        if not cc_match.empty:
            costcode_val = cc_match.iloc[0]["Costcode"]
            st.success(f"✅ Costcode: `{costcode_val}`")
        else:
            costcode_val = ""
            if cc_type != "— ประเภท —" or cc_work != "— Work —" or cc_loc != "— Location —":
                st.caption("ไม่พบ Costcode ที่ตรงกัน")
            else:
                st.caption("ยังไม่ได้เลือก Costcode")

        st.markdown("#### 📍 สถานที่ส่ง")
        location_choice = st.selectbox("สถานที่ส่ง", options=LOCATION_OPTIONS, key="location_choice")
        location_val    = "" if location_choice == "— ไม่ระบุ —" else location_choice

        # ── Export ──
        st.divider()
        export_disabled = not requester_name.strip()
        if export_disabled:
            st.warning("⚠️ กรุณาเลือกผู้เบิกก่อน")

        if st.button(
            "📥 สร้าง Form เบิกของ",
            type="primary",
            use_container_width=True,
            disabled=export_disabled,
        ):
            try:
                doc_number  = doc_number_input.strip() or get_next_doc_number()
                excel_bytes = export_to_form(
                    qty_data, doc_number, requester_name, employee_id,
                    costcode=costcode_val, location=location_val,
                )
                st.session_state.dl_bytes        = excel_bytes
                st.session_state.dl_fname        = f"{doc_number}.xlsx"
                st.session_state.dl_msg          = f"✅ {doc_number} — {requester_name} ({len(qty_data)} รายการ)"
                st.session_state.dl_drive_url    = None  # reset เพื่ออัปโหลดใหม่
                st.session_state.pending_history = {
                    "qty_data":       qty_data,
                    "doc_number":     doc_number,
                    "requester_name": requester_name,
                    "employee_id":    employee_id,
                }
                st.rerun()
            except Exception as e:
                st.error(f"เกิดข้อผิดพลาด: {e}")

        if st.session_state.dl_bytes is None:
            st.info("💡 กด 'สร้าง Form เบิกของ' แล้วปุ่มดาวน์โหลดจะปรากฏด้านบน")


# ══════════════════════════════════════════
# TAB 3 — ประวัติ
# ══════════════════════════════════════════
with tab3:
    col_hist_title, col_hist_refresh, col_hist_del = st.columns([4, 1, 1])
    with col_hist_title:
        st.subheader("📜 ประวัติการเบิก")
    with col_hist_refresh:
        st.write("")
        if st.button("🔄 รีเฟรช", use_container_width=True, help="โหลดประวัติล่าสุด"):
            st.rerun()
    with col_hist_del:
        st.write("")
        if st.button("🗑️ ลบทั้งหมด", use_container_width=True, type="secondary"):
            st.session_state["confirm_delete_hist"] = True

    if st.session_state.get("confirm_delete_hist"):
        st.warning("⚠️ ยืนยันลบประวัติทั้งหมด? กดยืนยันเพื่อดำเนินการ หรือกดยกเลิกเพื่อออก")
        col_yes, col_no = st.columns(2)
        with col_yes:
            if st.button("✅ ยืนยัน ลบเลย", use_container_width=True, type="primary"):
                delete_all_history()
                st.session_state["confirm_delete_hist"] = False
                st.success("ลบประวัติทั้งหมดแล้ว")
                st.rerun()
        with col_no:
            if st.button("❌ ยกเลิก", use_container_width=True):
                st.session_state["confirm_delete_hist"] = False
                st.rerun()

    hist = load_history()
    if hist.empty:
        st.info("ยังไม่มีประวัติการเบิก")
    else:
        st.dataframe(hist.reset_index(drop=True), use_container_width=True, hide_index=True)
        st.caption(f"ประวัติทั้งหมด: {len(hist):,} รายการ  |  กด 🔄 เพื่อดูข้อมูลล่าสุด")
        fname_hist = f"history_{datetime.now().strftime('%Y%m%d')}.xlsx"
        st.download_button(
            label    = "⬇️ ดาวน์โหลดประวัติ (.xlsx)",
            data     = history_to_excel_bytes(),
            file_name= fname_hist,
            mime     = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
